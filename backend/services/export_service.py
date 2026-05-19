import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import csv
import io
from datetime import datetime


class ExportService:
    @staticmethod
    def generate_excel(course, students, attendance_map, dates):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        ws['A1'] = f"Course: {course['name']}"
        ws['A2'] = f"Code: {course['code']}"
        ws['A3'] = f"Teacher: {course.get('teacherName','')}"
        ws['A4'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d')}"
        for r in range(1, 5):
            ws.cell(row=r, column=1).font = Font(bold=True)

        headers = ['Student Name', 'Roll Number'] + dates + ['Total Present', 'Total Classes', 'Percentage']
        header_fill = PatternFill(fill_type='solid', fgColor='2F4F4F')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        green_fill = PatternFill(fill_type='solid', fgColor='C6EFCE')
        red_fill = PatternFill(fill_type='solid', fgColor='FFC7CE')
        grey_fill = PatternFill(fill_type='solid', fgColor='F2F2F2')
        yellow_fill = PatternFill(fill_type='solid', fgColor='FFEB9C')

        for row, student in enumerate(students, 7):
            records = attendance_map.get(student['id'], {})
            present_count = 0
            ws.cell(row=row, column=1, value=student.get('name', ''))
            ws.cell(row=row, column=2, value=student.get('roll', ''))

            for col, date in enumerate(dates, 3):
                if date in records:
                    if records[date]:
                        cell = ws.cell(row=row, column=col, value='P')
                        cell.fill = green_fill
                        present_count += 1
                    else:
                        cell = ws.cell(row=row, column=col, value='A')
                        cell.fill = red_fill
                else:
                    cell = ws.cell(row=row, column=col, value='-')
                    cell.fill = grey_fill
                cell.alignment = Alignment(horizontal='center')

            total = len(dates)
            pct = round((present_count / total * 100), 1) if total > 0 else 0
            ws.cell(row=row, column=len(dates) + 3, value=present_count)
            ws.cell(row=row, column=len(dates) + 4, value=total)
            pct_cell = ws.cell(row=row, column=len(dates) + 5, value=f"{pct}%")
            if pct >= 75:
                pct_cell.fill = green_fill
            elif pct >= 50:
                pct_cell.fill = yellow_fill
            else:
                pct_cell.fill = red_fill

        # Auto-fit columns
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                v = '' if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[col_letter].width = max_len + 4

        ws.freeze_panes = 'C7'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_csv(course, students, attendance_map, dates):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([f"Course: {course['name']}"])
        writer.writerow([f"Code: {course['code']}"])
        writer.writerow([f"Teacher: {course.get('teacherName','')}"])
        writer.writerow([f"Export Date: {datetime.now().strftime('%Y-%m-%d')}"])
        writer.writerow([])
        writer.writerow(['Student Name', 'Roll Number'] + dates + ['Total Present', 'Total Classes', 'Percentage'])
        for student in students:
            records = attendance_map.get(student['id'], {})
            present_count = 0
            row = [student.get('name', ''), student.get('roll', '')]
            for date in dates:
                if date in records:
                    if records[date]:
                        row.append('P'); present_count += 1
                    else:
                        row.append('A')
                else:
                    row.append('-')
            total = len(dates)
            pct = round((present_count / total * 100), 1) if total > 0 else 0
            row += [present_count, total, f"{pct}%"]
            writer.writerow(row)
        buffer.seek(0)
        return buffer


def _format_session_label(date_str: str, timestamp_str: str) -> str:
    """Human-readable session header for export columns.

    Single session on a date → "2026-05-12".
    Multiple sessions same date → "2026-05-12 10:30", "2026-05-12 14:15", etc.
    Caller decides which form to use based on the per-date session count.
    """
    if not timestamp_str:
        return date_str or ""
    try:
        dt = datetime.fromisoformat(timestamp_str.replace("Z", "+00:00"))
        return f"{date_str} {dt.strftime('%H:%M')}"
    except Exception:
        return date_str or ""


def build_attendance_export_payload(course: dict, db):
    """Build (course, students_list, attendance_map, session_labels) for export.

    Each committed attendance session (one teacher Accept) gets its own
    column in the report. If the teacher ran multiple sessions on the same
    day, each becomes its own column tagged with date + start time —
    "2026-05-12 10:30", "2026-05-12 14:15", etc. Records without a
    sessionId (uncommitted bluetooth / code marks) are ignored.
    """
    course_id = course.get("id")
    students_list = list(course.get("enrolledStudents") or [])

    # First pass: gather every committed session across all students, recording
    # its date and the earliest timestamp (we use that as the session "start
    # time" for the column header).
    session_dates: dict[str, str] = {}     # sessionId -> date
    session_earliest: dict[str, str] = {}  # sessionId -> earliest ISO timestamp

    # Also build per-student per-session present/absent map.
    per_student_session: dict[str, dict[str, bool]] = {}

    for student in students_list:
        sid = (student or {}).get("id")
        if not sid:
            continue
        doc_id = f"{course_id}_{sid}"
        snap = db.collection("attendance").document(doc_id).get()
        student_map: dict[str, bool] = {}
        if snap.exists:
            data = snap.to_dict() or {}
            for rec in data.get("records", []) or []:
                session_id = rec.get("sessionId")
                if not session_id:
                    continue  # skip uncommitted records
                date = rec.get("date") or ""
                ts = rec.get("timestamp") or ""

                # Record session metadata once per session.
                if session_id not in session_dates:
                    session_dates[session_id] = date
                if session_id not in session_earliest or ts < session_earliest[session_id]:
                    session_earliest[session_id] = ts

                # Per-student per-session — present if ANY record for that
                # session is True (in practice manual submit is single-record).
                prior = student_map.get(session_id, False)
                student_map[session_id] = prior or bool(rec.get("present"))
        per_student_session[sid] = student_map

    # Sort sessions chronologically by their earliest timestamp.
    sorted_session_ids = sorted(
        session_dates.keys(),
        key=lambda s: (session_earliest.get(s, ""), session_dates.get(s, "")),
    )

    # Decide labels — only include the time component when a date has more than
    # one session, so single-session days stay clean.
    sessions_per_date: dict[str, int] = {}
    for sid in sorted_session_ids:
        date = session_dates[sid]
        sessions_per_date[date] = sessions_per_date.get(date, 0) + 1

    session_labels: list[str] = []
    label_by_session: dict[str, str] = {}
    for sid in sorted_session_ids:
        date = session_dates[sid]
        if sessions_per_date.get(date, 0) > 1:
            label = _format_session_label(date, session_earliest.get(sid, ""))
        else:
            label = date
        # Disambiguate collisions defensively (same date + time down to the minute)
        original = label
        suffix = 2
        while label in label_by_session.values():
            label = f"{original} #{suffix}"
            suffix += 1
        session_labels.append(label)
        label_by_session[sid] = label

    # Translate per-student per-session map into per-student per-label map.
    attendance_map: dict[str, dict[str, bool]] = {}
    for sid, student_map in per_student_session.items():
        attendance_map[sid] = {
            label_by_session[session_id]: present
            for session_id, present in student_map.items()
            if session_id in label_by_session
        }

    return course, students_list, attendance_map, session_labels
